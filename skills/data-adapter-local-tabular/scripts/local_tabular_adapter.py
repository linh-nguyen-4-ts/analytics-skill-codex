#!/usr/bin/env python3
"""Local adapter for CSV and XLSX files using DuckDB."""

from __future__ import annotations

import argparse
import csv
import json
import os
import tempfile
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

try:
    import duckdb
except ImportError:
    print(
        "Missing dependency. Install with: pip3 install duckdb openpyxl",
        file=os.sys.stderr,
    )
    raise SystemExit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    print(
        "Missing dependency. Install with: pip3 install openpyxl",
        file=os.sys.stderr,
    )
    raise SystemExit(1)


def split_source(source: str) -> tuple[Path, str | None]:
    path_text, sep, sheet = source.partition("#")
    path = Path(path_text).expanduser().resolve()
    if not path.exists():
        raise SystemExit(f"Source not found: {path}")
    return path, sheet or None


def detect_file_type(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix == ".xlsx":
        return "xlsx"
    raise SystemExit("Supported file types are .csv and .xlsx")


def quote_sql_string(value: str) -> str:
    escaped = value.replace("'", "''")
    return f"'{escaped}'"


def stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    return str(value)


def is_empty_row(row: tuple[object, ...]) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def normalize_headers(header_row: tuple[object, ...]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(header_row, start=1):
        base = stringify_cell(value).strip()
        if not base:
            base = f"col_{index}"
        base = base.replace("\n", " ").strip()
        count = counts.get(base, 0)
        counts[base] = count + 1
        header = base if count == 0 else f"{base}_{count + 1}"
        headers.append(header)
    return headers


def list_sheets(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def pick_sheet(path: Path, requested_sheet: str | None) -> str:
    sheets = list_sheets(path)
    if requested_sheet:
        if requested_sheet not in sheets:
            raise SystemExit(
                f"Sheet '{requested_sheet}' not found. Available sheets: {', '.join(sheets)}"
            )
        return requested_sheet
    if not sheets:
        raise SystemExit("Workbook has no sheets.")
    return sheets[0]


def export_sheet_to_csv(path: Path, sheet_name: str) -> tuple[str, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)

        header_row = None
        for row in rows:
            if not is_empty_row(tuple(row)):
                header_row = tuple(row)
                break
        if header_row is None:
            raise SystemExit(f"Sheet '{sheet_name}' is empty.")

        headers = normalize_headers(header_row)

        temp = tempfile.NamedTemporaryFile(
            mode="w",
            newline="",
            encoding="utf-8",
            suffix=".csv",
            delete=False,
        )
        with temp as handle:
            writer = csv.writer(handle)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([stringify_cell(value) for value in row])
        return temp.name, headers
    finally:
        workbook.close()


@contextmanager
def prepared_source(source: str) -> Iterator[dict[str, object]]:
    path, requested_sheet = split_source(source)
    file_type = detect_file_type(path)
    temp_csv = None
    sheet_name = None

    try:
        if file_type == "csv":
            csv_path = str(path)
        else:
            sheet_name = pick_sheet(path, requested_sheet)
            csv_path, _headers = export_sheet_to_csv(path, sheet_name)
            temp_csv = csv_path

        connection = duckdb.connect()
        csv_literal = quote_sql_string(csv_path)
        connection.execute(
            "CREATE OR REPLACE VIEW source_data AS "
            f"SELECT * FROM read_csv_auto({csv_literal}, sample_size=-1, header=true)"
        )
        yield {
            "connection": connection,
            "path": str(path),
            "file_type": file_type,
            "sheet_name": sheet_name,
        }
    finally:
        try:
            connection.close()  # type: ignore[name-defined]
        except Exception:
            pass
        if temp_csv and os.path.exists(temp_csv):
            os.unlink(temp_csv)


def serialize_row(row: dict) -> dict:
    output = {}
    for key, value in row.items():
        if hasattr(value, "isoformat") and not isinstance(value, str):
            output[key] = value.isoformat()
        else:
            output[key] = value
    return output


def fetch_rows(connection: duckdb.DuckDBPyConnection, sql: str, limit: int | None = None) -> list[dict]:
    query = sql
    if limit is not None:
        query = f"SELECT * FROM ({sql}) AS subquery LIMIT {limit}"
    cursor = connection.execute(query)
    columns = [item[0] for item in cursor.description]
    return [serialize_row(dict(zip(columns, row))) for row in cursor.fetchall()]


def cmd_inspect(args: argparse.Namespace) -> int:
    path, requested_sheet = split_source(args.source)
    file_type = detect_file_type(path)
    payload: dict[str, object] = {
        "path": str(path),
        "file_type": file_type,
        "size_bytes": path.stat().st_size,
    }

    if file_type == "xlsx":
        sheets = list_sheets(path)
        payload["sheets"] = sheets
        payload["default_sheet"] = sheets[0] if sheets else None
        payload["requested_sheet"] = requested_sheet

    with prepared_source(args.source) as ctx:
        rows = fetch_rows(ctx["connection"], "DESCRIBE source_data")
        payload["columns"] = rows
        if ctx["sheet_name"]:
            payload["selected_sheet"] = ctx["sheet_name"]

    print(json.dumps(payload, ensure_ascii=False, indent=2))
    return 0


def cmd_describe(args: argparse.Namespace) -> int:
    with prepared_source(args.source) as ctx:
        rows = fetch_rows(ctx["connection"], "DESCRIBE source_data")
        for row in rows:
            print(json.dumps(row, ensure_ascii=False))
    return 0


def cmd_preview(args: argparse.Namespace) -> int:
    with prepared_source(args.source) as ctx:
        rows = fetch_rows(ctx["connection"], "SELECT * FROM source_data", limit=args.limit)
        for row in rows:
            print(json.dumps(row, ensure_ascii=False))
    return 0


def quote_identifier(name: str) -> str:
    escaped = name.replace('"', '""')
    return f'"{escaped}"'


def cmd_profile(args: argparse.Namespace) -> int:
    with prepared_source(args.source) as ctx:
        connection = ctx["connection"]
        row_count = connection.execute("SELECT COUNT(*) FROM source_data").fetchone()[0]
        columns = fetch_rows(connection, "DESCRIBE source_data")

        summary = {
            "path": ctx["path"],
            "file_type": ctx["file_type"],
            "sheet_name": ctx["sheet_name"],
            "row_count": row_count,
            "columns": [],
        }

        for column in columns:
            column_name = column["column_name"]
            identifier = quote_identifier(column_name)
            null_count = connection.execute(
                f"SELECT COUNT(*) FROM source_data WHERE {identifier} IS NULL"
            ).fetchone()[0]
            distinct_count = connection.execute(
                f"SELECT COUNT(DISTINCT {identifier}) FROM source_data"
            ).fetchone()[0]
            sample_rows = fetch_rows(
                connection,
                (
                    f"SELECT DISTINCT {identifier} AS value "
                    f"FROM source_data WHERE {identifier} IS NOT NULL"
                ),
                limit=args.sample_values,
            )
            summary["columns"].append(
                {
                    "column_name": column_name,
                    "column_type": column["column_type"],
                    "null_count": null_count,
                    "null_rate": (null_count / row_count) if row_count else 0,
                    "distinct_count": distinct_count,
                    "sample_values": [row["value"] for row in sample_rows],
                }
            )

        print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


def load_sql(args: argparse.Namespace) -> str:
    if args.file:
        return Path(args.file).read_text()
    if args.sql:
        return args.sql
    raise SystemExit("Provide SQL directly or with --file.")


def cmd_query(args: argparse.Namespace) -> int:
    with prepared_source(args.source) as ctx:
        sql = load_sql(args).replace("{table}", "source_data")
        rows = fetch_rows(ctx["connection"], sql, limit=args.max_rows)

        if args.csv:
            if not rows:
                return 0
            writer = csv.DictWriter(os.sys.stdout, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
            return 0

        for row in rows:
            print(json.dumps(row, ensure_ascii=False))
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    inspect_parser = subparsers.add_parser("inspect")
    inspect_parser.add_argument("source")
    inspect_parser.set_defaults(func=cmd_inspect)

    describe_parser = subparsers.add_parser("describe")
    describe_parser.add_argument("source")
    describe_parser.set_defaults(func=cmd_describe)

    preview_parser = subparsers.add_parser("preview")
    preview_parser.add_argument("source")
    preview_parser.add_argument("--limit", type=int, default=5)
    preview_parser.set_defaults(func=cmd_preview)

    profile_parser = subparsers.add_parser("profile")
    profile_parser.add_argument("source")
    profile_parser.add_argument("--sample-values", type=int, default=5)
    profile_parser.set_defaults(func=cmd_profile)

    query_parser = subparsers.add_parser("query")
    query_parser.add_argument("source")
    query_parser.add_argument("sql", nargs="?")
    query_parser.add_argument("--file")
    query_parser.add_argument("--csv", action="store_true")
    query_parser.add_argument("--max-rows", type=int, default=500)
    query_parser.set_defaults(func=cmd_query)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    raise SystemExit(main())
